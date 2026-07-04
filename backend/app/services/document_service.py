from __future__ import annotations
import uuid
from pathlib import Path
from fastapi import UploadFile, HTTPException
from app.core.config import settings
from app.db.postgres import pg_conn
from app.schemas.contracts import DocumentForExtraction, PageText, SourceMetadata

MAX_UPLOAD_SIZE = 50 * 1024 * 1024
ALLOWED_SUFFIXES = {".pdf", ".docx", ".xlsx", ".xlsm", ".txt", ".md", ".csv"}


class DocumentService:
    def save_upload(self, file: UploadFile, access_level: str = "internal") -> dict:
        Path(settings.storage_dir).mkdir(parents=True, exist_ok=True)
        doc_id = f"doc_{uuid.uuid4().hex[:12]}"
        suffix = Path(file.filename or "uploaded.bin").suffix.lower()
        if suffix not in ALLOWED_SUFFIXES:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {suffix or 'no extension'}")
        storage_path = Path(settings.storage_dir) / f"{doc_id}{suffix}"
        content = file.file.read(MAX_UPLOAD_SIZE + 1)
        if len(content) > MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=413, detail="File too large; max 50 MB")
        storage_path.write_bytes(content)
        with pg_conn() as conn:
            conn.execute(
                """
                INSERT INTO documents (id, filename, content_type, storage_path, access_level, status)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (doc_id, file.filename, file.content_type, str(storage_path), access_level, "uploaded"),
            )
        return {"document_id": doc_id, "filename": file.filename, "access_level": access_level, "status": "uploaded"}

    def mark_indexed(self, document_id: str) -> None:
        with pg_conn() as conn:
            conn.execute(
                "UPDATE documents SET status = 'indexed' WHERE id = %s",
                (document_id,),
            )

    def mark_processed(self, document_id: str) -> None:
        with pg_conn() as conn:
            conn.execute(
                "UPDATE documents SET status = 'processed', processed_at = now() WHERE id = %s",
                (document_id,),
            )

    def get_document(self, document_id: str, include_storage_path: bool = False, visible_access_levels: list[str] | None = None) -> dict | None:
        visible_access_levels = visible_access_levels or ["public", "internal", "confidential"]
        with pg_conn() as conn:
            row = conn.execute(
                """
                SELECT id, filename, content_type, storage_path, access_level, status, created_at, processed_at
                FROM documents WHERE id = %s AND coalesce(access_level, 'internal') = ANY(%s)
                """,
                (document_id, visible_access_levels),
            ).fetchone()
        if not row:
            return None
        result = {
            "document_id": row[0],
            "filename": row[1],
            "content_type": row[2],
            "access_level": row[4] or "internal",
            "status": row[5],
            "created_at": row[6].isoformat() if row[6] else None,
            "processed_at": row[7].isoformat() if row[7] else None,
        }
        if include_storage_path:
            result["storage_path"] = row[3]
        return result

    def list_documents(self, include_storage_path: bool = False, visible_access_levels: list[str] | None = None) -> list[dict]:
        visible_access_levels = visible_access_levels or ["public", "internal", "confidential"]
        with pg_conn() as conn:
            rows = conn.execute(
                """
                SELECT id, filename, content_type, storage_path, access_level, status, created_at, processed_at
                FROM documents
                WHERE coalesce(access_level, 'internal') = ANY(%s)
                ORDER BY created_at DESC
                LIMIT 100
                """,
                (visible_access_levels,),
            ).fetchall()
        items = []
        for r in rows:
            item = {
                "document_id": r[0],
                "filename": r[1],
                "content_type": r[2],
                "access_level": r[4] or "internal",
                "status": r[5],
                "created_at": r[6].isoformat() if r[6] else None,
                "processed_at": r[7].isoformat() if r[7] else None,
            }
            if include_storage_path:
                item["storage_path"] = r[3]
            items.append(item)
        return items

    def build_extraction_payload(self, document_id: str, visible_access_levels: list[str] | None = None) -> DocumentForExtraction:
        doc = self.get_document(
            document_id,
            include_storage_path=True,
            visible_access_levels=visible_access_levels or ["public"],
        )
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found or access denied")

        path = Path(doc["storage_path"])
        pages = self.extract_text_pages(path, doc.get("content_type") or "")
        if not pages:
            pages = [
                PageText(
                    page=1,
                    text=(
                        f"[TEXT_EXTRACTION_WARNING] Backend could not extract text from uploaded file. "
                        f"Shared storage path for ML service: {doc['storage_path']}"
                    ),
                )
            ]

        return DocumentForExtraction(
            document_id=document_id,
            filename=doc["filename"],
            document_type="report",
            language_hint="ru",
            source_type="internal_report",
            access_level=doc.get("access_level") or "internal",
            metadata=SourceMetadata(title=doc["filename"]),
            pages=pages,
        )


    def build_mock_extraction_payload(self) -> DocumentForExtraction:
        """Explicit mock payload used only by /api/documents/process-mock.

        Important: uploaded document processing must never fall back to this method;
        otherwise a typo or inaccessible document_id would import synthetic facts.
        """
        return DocumentForExtraction(
            document_id="doc_mock_001",
            filename="doc_mock_001.mock",
            document_type="report",
            language_hint="ru",
            source_type="internal_report",
            access_level="internal",
            metadata=SourceMetadata(
                title="Анонимизированный R&D корпус: гидрометаллургия и экология",
                authors=["Иванов И.И.", "Petrov P."],
                year=2024,
                country="RU",
                organization="Demo R&D Lab",
            ),
            pages=[PageText(page=1, text="Explicit mock demo document. ML mock returns mock_extracted_document.json.")],
        )

    def extract_text_pages(self, path: Path, content_type: str = "") -> list[PageText]:
        suffix = path.suffix.lower()
        try:
            if suffix in {".txt", ".md", ".csv", ".log"} or content_type.startswith("text/"):
                text = path.read_text(encoding="utf-8", errors="replace")
                return [PageText(page=1, text=text[:200_000])]

            if suffix == ".pdf":
                from pypdf import PdfReader
                reader = PdfReader(str(path))
                pages: list[PageText] = []
                for idx, page in enumerate(reader.pages[:200], start=1):
                    pages.append(PageText(page=idx, text=(page.extract_text() or "")[:50_000]))
                return [p for p in pages if p.text.strip()]

            if suffix == ".docx":
                from docx import Document
                doc = Document(str(path))
                text = "\n".join(p.text for p in doc.paragraphs if p.text)
                return [PageText(page=1, text=text[:200_000])] if text.strip() else []

            if suffix in {".xlsx", ".xlsm"}:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), read_only=True, data_only=True)
                pages: list[PageText] = []
                for idx, ws in enumerate(wb.worksheets[:50], start=1):
                    lines = []
                    for row in ws.iter_rows(max_row=500, values_only=True):
                        values = [str(v) for v in row if v is not None]
                        if values:
                            lines.append("\t".join(values))
                    text = f"Sheet: {ws.title}\n" + "\n".join(lines)
                    if text.strip():
                        pages.append(PageText(page=idx, text=text[:50_000]))
                return pages
        except Exception as exc:
            return [PageText(page=1, text=f"[TEXT_EXTRACTION_ERROR] {type(exc).__name__}: {exc}")]
        return []
